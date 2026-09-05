#!/usr/bin/env python3
"""Parse every workbook in work/xlsx into nodes (one per sheet) with rows already
normalised by rules.py. Reads with data_only=True so the cached ΠΡΟΟΔΕΥΤΙΚΟ is
visible. Nothing here decides anything — it records, and it repairs only what
rules.py allows (year spikes, missing dates inherited from the line above), each
repair written into the row's note."""
import argparse, datetime as dt, json, os, warnings
from decimal import Decimal
import openpyxl
from rules import detect_header, classify, fix_date, raw_balance, is_num, d2, to_date, Unknown, FIELD_KEYS, norm
warnings.filterwarnings('ignore')

HERE = os.path.dirname(os.path.abspath(__file__))
WORK = os.path.join(HERE, 'work')
PICK = ('seq', 'date', 'date_end', 'route', 'advance', 'expenses', 'value', 'balance', 'running', 'cash', 'bank')
INHERIT_NOTE = 'ημ/νία από προηγούμενη γραμμή (κενή στο Excel)'

def jsonable(v):
    if isinstance(v, (dt.datetime, dt.date)): return v.isoformat()[:10]
    if isinstance(v, float) and v != v: return None
    return v

def add_note(entry, text):
    entry['note'] = (entry['note'] + ' · ' + text) if entry.get('note') else text

TEXT_AMOUNT_FIELDS = ('advance', 'expenses', 'value', 'balance')

def collect_text_amounts(cells, rn, out):
    # I8: a cell typed as text in an amount column (e.g. a formula error, or a
    # human typing "?") silently becomes 0 once openpyxl/is_num see it — the row
    # looks like a real amount and vanishes. Flag it instead of guessing.
    for field in TEXT_AMOUNT_FIELDS:
        v = cells.get(field)
        if isinstance(v, str) and v.strip():
            out.append({'row': rn, 'field': field, 'text': v[:40]})

def parse_sheet(ws, today):
    head = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 400), values_only=True))
    h = detect_header(head)
    if h is None: return None
    cols = h['cols']
    rows, unknown, cells_used = [], [], []
    pending = []  # leading undated rows waiting for the first dated row
    after_totals = False
    memos = []  # rows below ΣΥΝΟΛΟ line without a running value (loan reconciliations, notes)
    text_amount_rows = []
    totals_skipped = text_only = zero_net = 0
    for rn, raw in enumerate(ws.iter_rows(min_row=h['row'] + 1, values_only=True), h['row'] + 1):
        cells = {f: (raw[cols[f] - 1] if f in cols and cols[f] <= len(raw) else None) for f in PICK}
        cells['_row_text'] = ' '.join(str(v) for v in raw if isinstance(v, str))
        try:
            e = classify(cells)
        except Unknown as ex:
            unknown.append({'row': rn, 'reason': str(ex), 'cells': {k: jsonable(v) for k, v in cells.items() if k != '_row_text'}})
            if cells.get('date') is not None: cells_used.append(cells)   # a dated money line counts even if unclassified
            collect_text_amounts(cells, rn, text_amount_rows)
            continue
        if e == 'TOTALS': after_totals = True; totals_skipped += 1; continue
        if e is None:
            if any(v not in (None, '') for v in raw): text_only += 1
            continue
        # I8: only for rows that became a real entry (or ZERO_NET) — a repeated
        # header row mid-sheet ("ΕΛΑΒΕ"/"ΕΞΟΔΑ" as text in their own columns) is
        # already `None` above and correctly counted as text_only, not a typo to flag.
        collect_text_amounts(cells, rn, text_amount_rows)
        if e == 'ZERO_NET': zero_net += 1; continue
        # Below the ΣΥΝΟΛΟ line, what counts as a ledger entry depends on the sheet layout.
        # With a ΠΡΟΟΔΕΥΤΙΚΟ column, the running value decides: only entries with a running value
        # are ledger entries. Without one, the date decides: only dated rows are entries.
        # Undated rows (or rows without a running value) are memos (loan reconciliations, notes)
        # and must not move the balance. Kept for the owner report, never imported.
        if after_totals and (('running' in cols and not is_num(cells.get('running'))) or ('running' not in cols and to_date(cells.get('date')) is None)):
            memos.append({'row': rn, 'label': str(cells.get('route') or cells.get('cash') or '')[:80], 'amount': float(d2(cells.get('advance') if is_num(cells.get('advance')) else (cells.get('value') if is_num(cells.get('value')) else 0)))})
            continue
        inherited = False
        if e['entry_date'] is None:
            if not rows:
                # Leading undated row: hold it in pending for when the first dated row appears
                pending.append({'row': rn, 'cells': cells, 'entry': e})
                continue
            e['entry_date'] = rows[-1]['entry']['entry_date']; inherited = True; add_note(e, INHERIT_NOTE)
        else:
            # This is the first dated row: flush pending entries and inherit from it
            if pending:
                for p in pending:
                    p['entry']['entry_date'] = e['entry_date']; add_note(p['entry'], 'ημ/νία από επόμενη γραμμή (κενή στο Excel)')
                    rows.append({'row': p['row'], 'entry': p['entry'], 'cells': {k: jsonable(v) for k, v in p['cells'].items() if k != '_row_text'},
                                 'date_fix': None, 'date_problem': None, 'date_inherited': True})
                    cells_used.append(p['cells'])
                pending = []
        rows.append({'row': rn, 'entry': e, 'cells': {k: jsonable(v) for k, v in cells.items() if k != '_row_text'},
                     'date_fix': None, 'date_problem': None, 'date_inherited': inherited})
        cells_used.append(cells)
    # If pending still has entries at the end (a sheet with no dated row at all), push them to unknown
    for p in pending:
        unknown.append({'row': p['row'], 'reason': 'row without a date and no previous row to inherit from', 'cells': {k: jsonable(v) for k, v in p['cells'].items() if k != '_row_text'}})
    dates = [dt.date.fromisoformat(r['entry']['entry_date']) for r in rows]
    for i, r in enumerate(rows):
        nb = [d for d in dates[max(0, i - 3):i] + dates[i + 1:i + 4] if d <= today]
        fx = fix_date(dates[i], nb, today)
        if fx is None:
            r['date_problem'] = 'date %s is a spike and not repairable by year alone' % dates[i].isoformat()
        elif fx[1]:
            r['date_fix'] = {'from': dates[i].isoformat(), 'to': fx[0].isoformat(), 'note': fx[1]}
            r['entry']['entry_date'] = fx[0].isoformat(); add_note(r['entry'], fx[1])
        end = r['entry'].get('date_end')
        if end:
            start = dt.date.fromisoformat(r['entry']['entry_date']); e0 = dt.date.fromisoformat(end)
            # A return date before the departure OR way too far in the future is almost always a year typo;
            # repair it only when the year alone brings it to 0–60 days after departure.
            if e0 < start or e0 > start + dt.timedelta(days=60):
                cands = set()
                for y in {e0.year - 1, e0.year + 1, start.year, start.year + 1}:
                    try: cand = e0.replace(year=y)
                    except ValueError: continue
                    if start <= cand <= start + dt.timedelta(days=60): cands.add(cand)
                if len(cands) == 1:
                    fixed = cands.pop()
                    r['entry']['date_end'] = fixed.isoformat(); add_note(r['entry'], 'λήξη Excel %s → %s (έτος)' % (end, fixed.isoformat()))
                    r['date_fix'] = r['date_fix'] or {'from': end, 'to': fixed.isoformat(), 'note': 'λήξη: έτος'}
                else:
                    # Not a year typo (day/month slip or swapped cells). The balance does not
                    # depend on the return date, so the driver is not blocked: drop it, keep
                    # the original in the note and in date_fix.
                    if e0 < start:
                        msg = 'λήξη Excel %s μη έγκυρη (πριν την αναχώρηση), αφαιρέθηκε' % end
                    else:
                        msg = 'λήξη Excel %s μη έγκυρη (> 60 ημέρες μετά την αναχώρηση), αφαιρέθηκε' % end
                    add_note(r['entry'], msg)
                    r['date_fix'] = r['date_fix'] or {'from': end, 'to': None, 'note': msg}
                    r['entry']['date_end'] = None
    running_last, opening, breaks, consistent = None, None, [], None
    trailing, expected_final, residual = Decimal('0'), None, None
    if 'running' in cols:
        # Walk the cached ΠΡΟΟΔΕΥΤΙΚΟ against our own deltas. Rows without a running
        # value (a blank cell in the middle) accumulate into `acc` until the next
        # cached value. The first cached value fixes the opening balance; every later
        # jump that the row's own amounts do not explain is a break.
        prev_run, acc = None, Decimal('0')
        for r in rows:
            c = r['cells']
            delta = d2(c.get('value') if is_num(c.get('value')) else 0) - d2(c.get('advance') if is_num(c.get('advance')) else 0) + d2(c.get('expenses') if is_num(c.get('expenses')) else 0)
            acc += delta
            run = c.get('running')
            if not is_num(run): continue
            run = d2(run)
            if prev_run is None:
                opening = run - acc
            else:
                diff = run - (prev_run + acc)
                if abs(diff) > Decimal('0.05'):
                    breaks.append({'row': r['row'], 'entry_date': r['entry']['entry_date'], 'diff': str(diff)})
            prev_run, acc = run, Decimal('0')
            running_last = str(run)
        trailing = acc  # acc holds the deltas after the last cached running value
        if opening is not None and abs(opening) <= Decimal('0.05'): opening = None
        if running_last is not None:
            raw = raw_balance(cells_used)
            expected_final = str(d2(running_last) + trailing)
            tot = raw + (opening or Decimal('0')) + sum((Decimal(b['diff']) for b in breaks), Decimal('0'))
            gap = (d2(running_last) + trailing) - tot
            if abs(gap) <= Decimal('0.005'):
                consistent = True
                residual = None
            elif abs(gap) <= Decimal('1.00'):
                consistent = True
                residual = str(gap.quantize(Decimal('0.01')))
            else:
                consistent = False
                residual = None
    balance_sum = None
    if 'balance' in cols:
        vals = [c.get('balance') for c in cells_used if is_num(c.get('balance'))]
        balance_sum = str(sum((d2(v) for v in vals), d2(0))) if vals else None
    if expected_final is None:
        expected_final = balance_sum
    ds = sorted(dt.date.fromisoformat(r['entry']['entry_date']) for r in rows)
    return {'sheet': ws.title, 'header_row': h['row'], 'cols': cols, 'out_of_scope': h['out_of_scope'],
            'rows': rows, 'unknown': unknown, 'raw_final': str(raw_balance(cells_used)) if cells_used else None,
            'running_last': running_last, 'balance_sum': balance_sum,
            'opening_balance': str(opening) if opening is not None else None, 'running_breaks': breaks, 'running_consistent': consistent,
            'expected_final': expected_final, 'trailing_delta': str(trailing) if trailing != 0 else None, 'rounding_residual': residual,
            'first_date': ds[0].isoformat() if ds else None, 'last_date': ds[-1].isoformat() if ds else None,
            'n_rows': len(rows), 'totals_skipped': totals_skipped, 'text_only_skipped': text_only, 'zero_net_skipped': zero_net,
            'after_totals_skipped': len(memos), 'after_totals': memos, 'text_amount_rows': text_amount_rows}

def skip_reason(ws, today):
    # B1: a sheet that fails header detection used to vanish with no trace.
    # Record enough to tell "not a ledger sheet" apart from "the header row is
    # unusual" without opening the workbook by hand.
    head = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 400), values_only=True))
    non_empty = sum(1 for row in head for v in row if v not in (None, ''))
    blob = norm(' '.join(str(v) for row in head for v in row if isinstance(v, str)))
    matched = [field for field, keys in FIELD_KEYS if any(k in blob for k in keys)]
    return non_empty, matched

def main(today=None):
    if today is None: today = dt.date.today()
    index = json.load(open(os.path.join(WORK, 'drive-index.json'), encoding='utf-8'))
    nodes, skipped_sheets = [], []
    for it in index:
        wb = openpyxl.load_workbook(it['local'], data_only=True, read_only=True)
        for ws in wb.worksheets:
            n = parse_sheet(ws, today)
            if n is None:
                non_empty, matched = skip_reason(ws, today)
                skipped_sheets.append({'file_id': it['id'], 'file_name': it['name'], 'sheet': ws.title,
                                        'non_empty_cells': non_empty, 'matched_fields': matched})
                continue
            n.update({'file_id': it['id'], 'file_name': it['name'], 'path': it['path'],
                      'folder': 'ΣΤΑΜΑΤΗΣΑΝ' if it['path'].startswith('ΣΤΑΜΑΤΗΣΑΝ/') else 'root', 'modified': it['modified']})
            nodes.append(n)
    out = {'generated': dt.datetime.now().isoformat(timespec='seconds'), 'today': today.isoformat(), 'nodes': nodes, 'skipped_sheets': skipped_sheets}
    json.dump(out, open(os.path.join(WORK, 'inventory.json'), 'w', encoding='utf-8'), ensure_ascii=False)
    R = [r for n in nodes for r in n['rows']]
    print('nodes %d · rows %d · unknown rows %d · date fixes %d · date problems %d · date inherited %d · totals skipped %d · text-only %d · out_of_scope %d · running inconsistent %d · sheets with breaks %d · opening balances %d · rounding residuals %d · trailing %d · zero-net %d · after-totals memos %d · text amounts %d · sheets skipped %d (non-empty %d)'
          % (len(nodes), len(R), sum(len(n['unknown']) for n in nodes), sum(1 for r in R if r['date_fix']), sum(1 for r in R if r['date_problem']),
             sum(1 for r in R if r['date_inherited']), sum(n['totals_skipped'] for n in nodes), sum(n['text_only_skipped'] for n in nodes),
             sum(n['out_of_scope'] for n in nodes), sum(1 for n in nodes if n['running_consistent'] is False), sum(1 for n in nodes if n['running_breaks']), sum(1 for n in nodes if n['opening_balance']), sum(1 for n in nodes if n['rounding_residual']), sum(1 for n in nodes if n['trailing_delta']), sum(n['zero_net_skipped'] for n in nodes), sum(n['after_totals_skipped'] for n in nodes),
             sum(len(n['text_amount_rows']) for n in nodes), len(skipped_sheets), sum(s['non_empty_cells'] for s in skipped_sheets)))

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--today', type=dt.date.fromisoformat, default=None, help='YYYY-MM-DD, default today (reproducible spike repair)')
    a = ap.parse_args()
    main(a.today)
