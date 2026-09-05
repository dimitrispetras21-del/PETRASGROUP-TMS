#!/usr/bin/env python3
"""Import driver Excel ledger(s) (ΑΙΜΙΛΙΟΣ.xlsx layout) into the TMS ledger.

Dry run by default: parses, classifies, validates every row against the same
rules the Worker enforces, computes the final balance and compares it with
the workbook's last ΠΡΟΟΔΕΥΤΙΚΟ. Nothing is sent unless --commit AND both
checks pass. Unknown row shapes STOP that file's run — the tool never guesses.

Accepts one or more xlsx files, or a directory (all *.xlsx inside, sorted):

    python3 tools/import_driver_ledger.py ~/Drive/μισθοδοσία/ΑΙΜΙΛΙΟΣ.xlsx
    python3 tools/import_driver_ledger.py ~/Drive/μισθοδοσία/
    python3 tools/import_driver_ledger.py ~/Drive/μισθοδοσία/ΑΙΜΙΛΙΟΣ.xlsx --commit --token "$TMS_JWT"
"""
import argparse, glob, hashlib, json, os, re, sys, unicodedata, urllib.request, urllib.error
from decimal import Decimal, ROUND_HALF_UP
import openpyxl

PROXY = 'https://petras-tms-backend-staging.petrasgroup.workers.dev'
MAP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'driver-ledger-map.json')
HEADER_ROW = 3
ISO_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')
DL_TYPES = ['trip', 'payment_cash', 'payment_bank', 'adjustment']
SUFFIX_RE = re.compile(r'\s*\(\d+\)\s*$')

def d2(v):
    return Decimal(str(v or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def d2_or_none(v):
    """Same rounding as d2, but an empty Excel cell stays NULL. The DB and the
    UI both treat NULL as "not entered yet" and 0 as "entered, zero" (migration
    011 comment: trip_value/advance/expenses NULL = pending, never 0 for
    unknown) — `v or 0` in d2() collapses that distinction, so callers that
    must preserve it (classify's trip fields) use this instead."""
    return None if v is None else float(d2(v))

def iso(dtv):
    return dtv.date().isoformat() if hasattr(dtv, 'date') else None

def classify(b, c, d, e, f, g, i):
    """One Excel row -> ledger entry dict, or None for a blank row.
    b=Α/Α c=start d=end e=description f=ΕΛΑΒΕ g=ΕΞΟΔΑ i=ΑΞΙΑ."""
    desc = (e or '').strip().upper()
    if not desc and not c and not d:
        return None
    start = c or d
    if start is None:
        raise ValueError(f'row without a date: {e!r}')
    value = d2(i)  # branch decision only — an empty ΑΞΙΑ cell still means "no value yet", not "not a trip"
    if b is not None or value > 0:
        return {'entry_type': 'trip', 'entry_date': iso(start), 'date_end': iso(d) if d else None,
                'route': (e or '').strip(), 'trip_value': d2_or_none(i), 'advance': d2_or_none(f), 'expenses': d2_or_none(g)}
    if 'ΜΕΤΡΗΤ' in desc:
        return {'entry_type': 'payment_cash', 'entry_date': iso(start), 'amount': float(d2(f))}
    if 'ΚΑΤΑΘΕΣ' in desc or 'ΤΡΑΠΕΖ' in desc:
        return {'entry_type': 'payment_bank', 'entry_date': iso(start), 'amount': float(d2(f))}
    raise ValueError(f'unknown row shape: {e!r} (F={f}, I={i}) — decide by hand')

def check_header_row(ws):
    """HEADER_ROW is hardcoded — a shifted layout used to silently scan from
    the wrong offset and lose the first real data row. Refuse to guess when
    row 3 doesn't look like the real header."""
    cells = ' '.join(str(ws.cell(HEADER_ROW, c).value or '') for c in range(1, ws.max_column + 1))
    if 'ΗΜΕΡΟΜΗΝΙΑ' not in cells or 'ΑΞΙΑ' not in cells:
        raise ValueError('κεφαλίδα δεν βρέθηκε στη γραμμή 3')

def other_sheets_with_data(wb):
    """Only worksheets[0] is parsed (HEADER_ROW math assumes it); a second
    sheet with real content used to be dropped in total silence. Report it,
    don't fail — some exports carry an empty extra tab that must not block
    the import."""
    warnings = []
    for ws in wb.worksheets[1:]:
        non_empty = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ''))
        if non_empty > 1:
            warnings.append(f'φύλλο {ws.title} έχει δεδομένα — αγνοείται')
    return warnings

def parse_workbook(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    check_header_row(ws)
    rows, anomalies = [], other_sheets_with_data(wb)
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        b, c, d, e, f, g, i = (ws.cell(r, col).value for col in (2, 3, 4, 5, 6, 7, 9))
        if isinstance(c, str) and c.strip().upper() == 'ΣΥΝΟΛΟ':
            break
        entry = classify(b, c, d, e, f, g, i)
        if entry is None:
            continue
        if c and d and c > d:
            anomalies.append(f'row {r}: C > D ({c.date()} > {d.date()}) — year typo? kept as is')
        if entry['entry_type'] == 'trip' and d2(f) == 0:
            anomalies.append(f'row {r}: trip with ΕΛΑΒΕ = 0 (allowed, reported)')
        entry['_row'] = r
        rows.append(entry)
    # cached value of the last ΠΡΟΟΔΕΥΤΙΚΟ (column K); None when the file was
    # never recalculated (e.g. written by openpyxl)
    ws2 = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    excel_final = None
    for r in range(ws2.max_row, HEADER_ROW, -1):
        v = ws2.cell(r, 11).value
        if isinstance(v, (int, float)):
            excel_final = d2(v); break
    return rows, anomalies, excel_final

def validate_row(row):
    """Mirror worker/src/ledger-rules.mjs validateNewEntry for the fields this
    importer emits — run before --commit so a row the Worker would refuse is
    caught here first, against the real Excel row number (row['_row']), not a
    payload array index. Messages match the Worker's wording exactly."""
    if row.get('entry_type') not in DL_TYPES:
        return 'entry_type must be one of ' + '|'.join(DL_TYPES)
    entry_date = row.get('entry_date')
    if not entry_date or not ISO_RE.match(entry_date):
        return 'entry_date must be YYYY-MM-DD'
    if row['entry_type'] == 'trip':
        date_end = row.get('date_end')
        if date_end is not None:
            if not ISO_RE.match(date_end):
                return 'date_end must be YYYY-MM-DD'
            if date_end < entry_date:
                return 'date_end must not be before entry_date'
        for f in ('trip_value', 'advance', 'expenses'):
            v = row.get(f)
            if v is not None and v < 0:
                return f + ' must be a number ≥ 0'
        route = (row.get('route') or '').strip()
        if not route and row.get('rt_id') is None:
            return 'route or rt_id required for a trip'
    else:
        amount = row.get('amount')
        if amount is None:
            return 'amount required'
        is_adjustment = row['entry_type'] == 'adjustment'
        if is_adjustment:
            if amount == 0:
                return 'amount must be ≠ 0'
        elif amount <= 0:
            return 'amount must be > 0'
    return None

def payload_rows(rows):
    """Filter out private (_*), None-valued, and empty-string keys from entry dicts.
    A trip with no description parses to route='' — sending that would store an
    empty string instead of leaving the column unset (Worker distinguishes the two).
    A None trip_value/advance/expenses (empty Excel cell) is dropped the same way,
    leaving the column NULL instead of writing a false 0 (see d2_or_none)."""
    return [{k: v for k, v in e.items() if not k.startswith('_') and v is not None and v != ''} for e in rows]

def compute_balance(rows):
    bal = Decimal('0')
    for e in rows:
        if e['entry_type'] == 'trip':
            bal += d2(e['trip_value']) - (d2(e['advance']) - d2(e['expenses']))
        else:
            bal -= d2(e['amount'])
    return bal.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def map_key(name):
    """Filename -> driver-ledger-map.json key. Greek str.upper() keeps accents
    ('Αιμίλιος' -> 'ΑΙΜΙΛΙΟΣ' with tonos on Ι) but the map's keys are typed
    without them — NFD-decompose and drop combining marks first, then strip a
    Drive-style duplicate suffix like ' (1)'."""
    name = SUFFIX_RE.sub('', name.strip())
    decomposed = unicodedata.normalize('NFD', name)
    return ''.join(c for c in decomposed if not unicodedata.combining(c)).strip().upper()

def resolve_paths(paths):
    """Expand directory arguments to their *.xlsx files, sorted; keep file
    arguments as given, in the order they were passed."""
    files = []
    for p in paths:
        if os.path.isdir(p):
            files.extend(sorted(glob.glob(os.path.join(p, '*.xlsx'))))
        else:
            files.append(p)
    return files

def post_import(payload, token):
    """POST to /costs/ledger/import and return parsed response dict.
    On HTTPError: calls sys.exit with server error body.
    On URLError: calls sys.exit with network error reason.
    """
    req = urllib.request.Request(PROXY + '/costs/ledger/import', data=json.dumps(payload).encode(),
                                 headers={'Content-Type': 'application/json', 'Authorization': 'Bearer ' + token}, method='POST')
    try:
        with urllib.request.urlopen(req) as res:
            return json.load(res)
    # HTTPError first: it subclasses URLError and carries the server's error body.
    except urllib.error.HTTPError as err:
        sys.exit(f'✗ HTTP {err.code}: {err.read().decode()[:300]}')
    except urllib.error.URLError as err:
        # Network error: DNS failure, connection refused, timeout, etc.
        sys.exit(f'✗ network error: {err.reason}')

def run_one(path, mapping, commit, token):
    """Process one workbook end to end. Never raises for an expected failure
    (bad header, unmapped driver, a row the Worker would refuse, a balance
    mismatch, a server error) — every one of those becomes
    {'ok': False, 'message': ...} so a batch run can report it and move on
    (dry run) or stop cleanly (--commit, see main())."""
    name = map_key(os.path.splitext(os.path.basename(path))[0])
    result = {'name': name, 'path': path, 'ok': False, 'message': '', 'anomalies': [], 'errors': [], 'committed': False}
    if name not in mapping:
        result['message'] = f'no driver id in driver-ledger-map.json for {name!r}'
        return result
    driver_id = mapping[name]
    try:
        rows, anomalies, excel_final = parse_workbook(path)
    except ValueError as err:
        result['message'] = str(err)
        return result
    result['anomalies'] = anomalies
    row_errors = [(e['_row'], validate_row(e)) for e in rows]
    row_errors = [(r, msg) for r, msg in row_errors if msg]
    if row_errors:
        result['errors'] = row_errors
        r0, msg0 = row_errors[0]
        extra = f' (+{len(row_errors) - 1} ακόμη γραμμή/ές)' if len(row_errors) > 1 else ''
        result['message'] = f'γραμμή {r0}: {msg0}{extra}'
        return result
    bal = compute_balance(rows)
    if excel_final is None or bal != excel_final:
        result['message'] = f'balance {bal} != Excel {excel_final}'
        return result
    result['ok'] = True
    result['driver_id'] = driver_id
    result['count'] = len(rows)
    result['balance'] = bal
    result['rows'] = rows
    if not commit:
        result['message'] = f'{driver_id} · {len(rows)} γραμμές · {bal} ✓'
        return result
    if not token:
        result['ok'] = False
        result['message'] = '--token (or $TMS_JWT) required for --commit'
        return result
    payload = {'driver_id': driver_id, 'file_name': os.path.basename(path),
               'file_hash': hashlib.sha256(open(path, 'rb').read()).hexdigest(),
               'rows': payload_rows(rows)}
    try:
        out = post_import(payload, token)
    except SystemExit as err:
        result['ok'] = False
        result['message'] = str(err)
        return result
    if d2(out['balance']) != bal:
        result['ok'] = False
        result['message'] = f'SERVER BALANCE DIFFERS ({out["balance"]} != {bal}) — batch {out.get("batch")}, ask the owner to run dl_cancel_batch'
        return result
    result['committed'] = True
    result['message'] = f'{driver_id} · {len(rows)} γραμμές · imported batch {out["batch"]}, balance {out["balance"]}'
    return result

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('paths', nargs='+', help='one or more .xlsx files, or a directory containing them')
    ap.add_argument('--commit', action='store_true')
    ap.add_argument('--token', default=os.environ.get('TMS_JWT'))
    a = ap.parse_args()
    files = resolve_paths(a.paths)
    if not files:
        sys.exit('✗ no .xlsx files found')
    mapping = json.load(open(MAP_FILE, encoding='utf-8'))
    ready = failed = 0
    for path in files:
        r = run_one(path, mapping, a.commit, a.token)
        for x in r['anomalies']:
            print(f'  ⚠ {x}')
        for row_no, msg in r['errors']:
            print(f'  ✗ γραμμή {row_no}: {msg}')
        if r['ok']:
            ready += 1
            print(f'✓ {r["name"]} → {r["message"]}')
        else:
            failed += 1
            print(f'✗ {r["name"]}: {r["message"]}')
            if a.commit:
                # Files before this one are already committed (each is its own atomic
                # RPC call) — continuing past a failure would import out of order and
                # make a partial batch harder to reason about than stopping cleanly.
                break
    untried = len(files) - ready - failed
    footer = f'{len(files)} αρχεία · {ready} έτοιμα · {failed} με σφάλμα'
    if untried:
        footer += f' · {untried} δεν δοκιμάστηκαν (διακοπή μετά το πρώτο σφάλμα· τα προηγούμενα ήδη εισήχθησαν)'
    print(footer)
    if failed or untried:
        sys.exit(1)

if __name__ == '__main__':
    main()
