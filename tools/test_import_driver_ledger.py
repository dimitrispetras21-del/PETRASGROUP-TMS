import unittest, tempfile, os, datetime as dt, sys, io, urllib.error
from decimal import Decimal
from unittest.mock import patch
import openpyxl
from import_driver_ledger import (parse_workbook, compute_balance, classify, payload_rows, post_import,
                                   validate_row, map_key, resolve_paths, run_one, HEADER_ROW)

def make_xlsx(rows):
    wb = openpyxl.Workbook(); ws = wb.active
    ws['E1'] = 'ΟΔΗΓΟΣ  ΤΕΣΤ'
    for col, h in zip('CEFGHIJK', ['ΗΜΕΡΟΜΗΝΙΑ', 'ΔΡΟΜΟΛΟΓΙΟ', 'ΕΛΑΒΕ', 'ΕΞΟΔΑ', 'ΥΠΟΛΟΙΠΟ', 'ΑΞΙΑ ΔΡΟΜΟΛΟΓΙΟΥ', 'ΥΠΟΛΟΙΠΟ', 'ΠΡΟΟΔΕΥΤΙΚΟ ΥΠΟΛΟΙΠΟ']):
        ws[f'{col}3'] = h
    r = 4
    for row in rows:
        for col, v in row.items(): ws[f'{col}{r}'] = v
        r += 1
    ws[f'C{r}'] = 'ΣΥΝΟΛΟ'; ws[f'F{r}'] = f'=SUM(F4:F{r-1})'
    p = os.path.join(tempfile.mkdtemp(), 't.xlsx'); wb.save(p); return p

class PayloadRowsTests(unittest.TestCase):
    def test_payload_rows_filters_underscore_and_none(self):
        rows = [{'entry_type':'trip','entry_date':'2024-08-10','date_end':None,'route':'X','trip_value':800.0,'advance':200.0,'expenses':0.0,'_row':4}]
        result = payload_rows(rows)
        expected = [{'entry_type':'trip','entry_date':'2024-08-10','route':'X','trip_value':800.0,'advance':200.0,'expenses':0.0}]
        self.assertEqual(result, expected)

    def test_payload_rows_filters_empty_string(self):
        # A trip with no description classifies with route='' — must not be sent
        # as an empty string (would store route='' instead of leaving it unset).
        rows = [{'entry_type':'trip','entry_date':'2024-08-10','route':'','trip_value':800.0,'advance':200.0,'expenses':0.0,'_row':5}]
        result = payload_rows(rows)
        expected = [{'entry_type':'trip','entry_date':'2024-08-10','trip_value':800.0,'advance':200.0,'expenses':0.0}]
        self.assertEqual(result, expected)

class ClassifyTests(unittest.TestCase):
    def test_trip(self):
        r = classify(1, dt.datetime(2024, 8, 10), dt.datetime(2024, 8, 15), 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 200, 0, 800)
        self.assertEqual(r, {'entry_type': 'trip', 'entry_date': '2024-08-10', 'date_end': '2024-08-15',
                             'route': 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 'trip_value': 800.0, 'advance': 200.0, 'expenses': 0.0})
    def test_cash_and_bank(self):
        self.assertEqual(classify(None, dt.datetime(2024, 9, 10), None, 'ΜΕΤΡΗΤΑ', 650, 0, 0),
                         {'entry_type': 'payment_cash', 'entry_date': '2024-09-10', 'amount': 650.0})
        self.assertEqual(classify(None, dt.datetime(2024, 9, 2), None, 'ΚΑΤΑΘΕΣΗ ΤΡΑΠΕΖΑ ETE', 600, 0, 0)['entry_type'], 'payment_bank')
    def test_unknown_row_raises(self):
        with self.assertRaises(ValueError):
            classify(None, dt.datetime(2024, 9, 2), None, 'ΔΩΡΟ ΠΑΣΧΑ', 100, 0, 0)
    def test_national_without_number_is_trip(self):
        self.assertEqual(classify(None, dt.datetime(2025, 8, 19), None, 'ΑΘΗΝΑ', 100, 0, 230)['entry_type'], 'trip')

class ParseTests(unittest.TestCase):
    def test_balance_and_year_typo_flag(self):
        p = make_xlsx([
            {'B': 1, 'C': dt.datetime(2024, 8, 10), 'D': dt.datetime(2024, 8, 15), 'E': 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 'F': 200, 'G': 0, 'I': 800},
            {'C': dt.datetime(2024, 8, 19), 'E': 'ΜΕΤΡΗΤΑ', 'F': 600, 'G': 0, 'I': 0},
            {'B': 2, 'C': dt.datetime(2025, 12, 27), 'D': dt.datetime(2025, 1, 3), 'E': 'ΒΕΡΟΙΑ-ΑΥΣΤΡΙΑ', 'F': 300, 'G': 88, 'I': 950},
        ])
        rows, anomalies, excel_final = parse_workbook(p)
        self.assertEqual(len(rows), 3)
        self.assertEqual(compute_balance(rows), Decimal('738.00'))   # (800-200) - 600 + (950-(300-88)) = 738
        self.assertTrue(any('C > D' in a for a in anomalies))
        self.assertIsNone(excel_final)   # openpyxl-written file has no cached formula values

class PostImportTests(unittest.TestCase):
    def test_http_error_caught_first_with_server_body(self):
        """HTTPError is caught before URLError and includes server response body."""
        error_body = b'{"error":"row 3: amount required"}'
        http_err = urllib.error.HTTPError(url='http://test', code=422, msg='Unprocessable', hdrs=None, fp=io.BytesIO(error_body))
        with patch('import_driver_ledger.urllib.request.urlopen', side_effect=http_err):
            with self.assertRaises(SystemExit) as cm:
                post_import({'driver_id': 'rec123', 'rows': []}, 'token_xyz')
            msg = str(cm.exception)
            self.assertTrue(msg.startswith('✗ HTTP 422:'), f'Got: {msg}')
            self.assertIn('amount required', msg)

    def test_url_error_caught_after_http_error(self):
        """URLError (non-HTTP) is caught and includes reason."""
        url_err = urllib.error.URLError('Name or service not known')
        with patch('import_driver_ledger.urllib.request.urlopen', side_effect=url_err):
            with self.assertRaises(SystemExit) as cm:
                post_import({'driver_id': 'rec123', 'rows': []}, 'token_xyz')
            msg = str(cm.exception)
            self.assertTrue(msg.startswith('✗ network error:'), f'Got: {msg}')
            self.assertIn('Name or service not known', msg)

class EmptyCellIsNoneTests(unittest.TestCase):
    """Finding (verify-db.md, «⚠ Important — NULL→0 στα κενά κελιά Excel»):
    an empty Excel cell in I/F/G must become NULL, not 0 — the ledger's own
    invariant is "NULL = pending, never 0 for unknown" (migration 011)."""
    def test_blank_advance_cell_is_none_not_zero(self):
        r = classify(1, dt.datetime(2024, 8, 10), None, 'ΒΕΡΟΙΑ', None, 0, 800)
        self.assertIsNone(r['advance'])
        self.assertEqual(r['expenses'], 0.0)
        self.assertEqual(r['trip_value'], 800.0)

    def test_blank_trip_value_cell_is_none_not_zero(self):
        r = classify(1, dt.datetime(2024, 8, 10), None, 'ΒΕΡΟΙΑ', 200, 0, None)
        self.assertIsNone(r['trip_value'])

    def test_literal_zero_stays_zero(self):
        r = classify(1, dt.datetime(2024, 8, 10), None, 'ΒΕΡΟΙΑ', 0, 0, 800)
        self.assertEqual(r['advance'], 0.0)

    def test_none_advance_is_absent_from_payload_zero_is_present(self):
        blank = classify(1, dt.datetime(2024, 8, 10), None, 'ΒΕΡΟΙΑ', None, 0, 800)
        blank['_row'] = 4
        literal_zero = classify(1, dt.datetime(2024, 8, 10), None, 'ΒΕΡΟΙΑ', 0, 0, 800)
        literal_zero['_row'] = 5
        p_blank, p_zero = payload_rows([blank, literal_zero])
        self.assertNotIn('advance', p_blank)
        self.assertIn('advance', p_zero)
        self.assertEqual(p_zero['advance'], 0.0)


class ValidateRowTests(unittest.TestCase):
    """validate_row mirrors worker/src/ledger-rules.mjs validateNewEntry for
    the fields the importer emits (verify-importer.md §3/§5, findings γ/δ/η/θ)."""
    def test_year_typo_row_rejected_date_end_before_entry_date(self):
        row = classify(2, dt.datetime(2025, 12, 27), dt.datetime(2025, 1, 3), 'ΒΕΡΟΙΑ-ΑΥΣΤΡΙΑ', 300, 88, 950)
        self.assertEqual(validate_row(row), 'date_end must not be before entry_date')

    def test_text_date_rejected_never_silently_none(self):
        # A text-formatted Excel date ('10/8/2024' as a str, not a datetime)
        # makes iso() return None silently (finding η) — validate_row must
        # catch the resulting missing entry_date as a loud, named error.
        row = classify(1, '10/8/2024', None, 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ', 200, 0, 800)
        self.assertIsNone(row['entry_date'])
        self.assertEqual(validate_row(row), 'entry_date must be YYYY-MM-DD')

    def test_alfa_with_blank_description_rejected_route_or_rt_id(self):
        # Α/Α present (so it classifies as a trip) but E (description) blank.
        row = classify(3, dt.datetime(2024, 8, 10), None, '', 200, 0, 800)
        self.assertEqual(validate_row(row), 'route or rt_id required for a trip')

    def test_negative_value_rejected(self):
        row = classify(4, dt.datetime(2024, 8, 10), None, 'ΒΕΡΟΙΑ', 200, 0, -50)
        self.assertEqual(validate_row(row), 'trip_value must be a number ≥ 0')

    def test_valid_trip_and_payment_pass(self):
        trip = classify(1, dt.datetime(2024, 8, 10), dt.datetime(2024, 8, 15), 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 200, 0, 800)
        payment = classify(None, dt.datetime(2024, 9, 10), None, 'ΜΕΤΡΗΤΑ', 650, 0, 0)
        self.assertIsNone(validate_row(trip))
        self.assertIsNone(validate_row(payment))

    def test_dry_run_stops_before_balance_check_real_pattern(self):
        # Same shape as the real ΑΙΜΙΛΙΟΣ.xlsx row 36 — the local "balance
        # matches" check can't see this because it doesn't look at entry_date.
        p = make_xlsx([
            {'B': 1, 'C': dt.datetime(2024, 8, 10), 'D': dt.datetime(2024, 8, 15), 'E': 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 'F': 200, 'G': 0, 'I': 800},
            {'B': 2, 'C': dt.datetime(2025, 12, 27), 'D': dt.datetime(2025, 1, 3), 'E': 'ΒΕΡΟΙΑ-ΑΥΣΤΡΙΑ', 'F': 300, 'G': 88, 'I': 950},
        ])
        rows, anomalies, excel_final = parse_workbook(p)
        errors = [(e['_row'], validate_row(e)) for e in rows]
        errors = [(r, msg) for r, msg in errors if msg]
        self.assertEqual(len(errors), 1)
        row_no, msg = errors[0]
        self.assertEqual(row_no, 5)  # header row 3, first data row 4, this is the 2nd data row
        self.assertEqual(msg, 'date_end must not be before entry_date')


class HeaderAndSheetsGuardTests(unittest.TestCase):
    def test_header_not_on_row_3_raises(self):
        wb = openpyxl.Workbook(); ws = wb.active
        ws['C2'] = 'ΗΜΕΡΟΜΗΝΙΑ'; ws['I2'] = 'ΑΞΙΑ'  # header one row too high
        ws['C4'] = dt.datetime(2024, 8, 10); ws['I4'] = 800
        p = os.path.join(tempfile.mkdtemp(), 't.xlsx'); wb.save(p)
        with self.assertRaises(ValueError) as cm:
            parse_workbook(p)
        self.assertIn('κεφαλίδα', str(cm.exception))

    def test_header_missing_axia_raises(self):
        wb = openpyxl.Workbook(); ws = wb.active
        ws['C3'] = 'ΗΜΕΡΟΜΗΝΙΑ'  # no ΑΞΙΑ anywhere on row 3
        p = os.path.join(tempfile.mkdtemp(), 't.xlsx'); wb.save(p)
        with self.assertRaises(ValueError):
            parse_workbook(p)

    def test_second_sheet_with_data_is_reported_not_fatal(self):
        p = make_xlsx([
            {'B': 1, 'C': dt.datetime(2024, 8, 10), 'D': dt.datetime(2024, 8, 15), 'E': 'ΒΕΡΟΙΑ', 'F': 200, 'G': 0, 'I': 800},
        ])
        wb = openpyxl.load_workbook(p)
        extra = wb.create_sheet('Πρόχειρο')
        extra['A1'] = 'σημείωση'; extra['A2'] = 'δεύτερο κελί'
        wb.save(p)
        rows, anomalies, excel_final = parse_workbook(p)
        self.assertEqual(len(rows), 1)  # sheet 1 data ignored for parsing...
        self.assertTrue(any('Πρόχειρο' in a and 'αγνοείται' in a for a in anomalies))  # ...but reported

    def test_second_sheet_with_one_cell_is_not_reported(self):
        p = make_xlsx([
            {'B': 1, 'C': dt.datetime(2024, 8, 10), 'D': dt.datetime(2024, 8, 15), 'E': 'ΒΕΡΟΙΑ', 'F': 200, 'G': 0, 'I': 800},
        ])
        wb = openpyxl.load_workbook(p)
        extra = wb.create_sheet('Κενό')
        extra['A1'] = 'μόνο ένα κελί'
        wb.save(p)
        rows, anomalies, excel_final = parse_workbook(p)
        self.assertFalse(any('Κενό' in a for a in anomalies))


class MapKeyTests(unittest.TestCase):
    def test_strips_accents_and_duplicate_suffix(self):
        self.assertEqual(map_key('Αιμίλιος (1)'), 'ΑΙΜΙΛΙΟΣ')

    def test_already_normalised_key_is_idempotent(self):
        self.assertEqual(map_key('ΑΙΜΙΛΙΟΣ'), 'ΑΙΜΙΛΙΟΣ')

    def test_lowercase_with_tonos(self):
        self.assertEqual(map_key('αιμίλιος'), 'ΑΙΜΙΛΙΟΣ')


class BatchSummaryTests(unittest.TestCase):
    """run_one() is what main()'s batch loop calls per file — test its
    per-file result directly rather than capturing stdout (verify-importer.md
    §6 "καμία μαζική εκτέλεση")."""
    def _rename(self, path, stem):
        # make_xlsx() always names the file 't.xlsx' — two files in one test
        # need distinct stems so they map to distinct driver-ledger-map keys.
        new_path = os.path.join(os.path.dirname(path), stem + '.xlsx')
        os.rename(path, new_path)
        return new_path

    def _good_file(self):
        p = make_xlsx([
            {'B': 1, 'C': dt.datetime(2024, 8, 10), 'D': dt.datetime(2024, 8, 15), 'E': 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 'F': 200, 'G': 0, 'I': 800},
        ])
        wb = openpyxl.load_workbook(p); ws = wb.active
        ws.cell(ws.max_row, 11, 600.0)  # cache ΠΡΟΟΔΕΥΤΙΚΟ = 800-200 = 600, matching compute_balance
        wb.save(p)
        return self._rename(p, 'ΚΑΛΟΣ')

    def _bad_file(self):
        p = make_xlsx([
            {'B': 1, 'C': dt.datetime(2024, 8, 10), 'D': dt.datetime(2024, 8, 15), 'E': 'ΒΕΡΟΙΑ-ΙΤΑΛΙΑ-ΒΕΡΟΙΑ', 'F': 200, 'G': 0, 'I': 800},
            {'B': 2, 'C': dt.datetime(2025, 12, 27), 'D': dt.datetime(2025, 1, 3), 'E': 'ΒΕΡΟΙΑ-ΑΥΣΤΡΙΑ', 'F': 300, 'G': 88, 'I': 950},
        ])
        return self._rename(p, 'ΚΑΚΟΣ')

    def test_good_file_ok_and_summary_line(self):
        good = self._good_file()
        mapping = {map_key(os.path.splitext(os.path.basename(good))[0]): 46}
        r = run_one(good, mapping, commit=False, token=None)
        self.assertTrue(r['ok'])
        self.assertEqual(r['message'], '46 · 1 γραμμές · 600.00 ✓')

    def test_bad_file_not_ok_names_row_and_reason(self):
        bad = self._bad_file()
        mapping = {map_key(os.path.splitext(os.path.basename(bad))[0]): 47}
        r = run_one(bad, mapping, commit=False, token=None)
        self.assertFalse(r['ok'])
        self.assertEqual(len(r['errors']), 1)
        self.assertIn('date_end must not be before entry_date', r['message'])

    def test_footer_counts_ready_and_failed_across_two_files(self):
        good, bad = self._good_file(), self._bad_file()
        mapping = {map_key(os.path.splitext(os.path.basename(good))[0]): 46,
                   map_key(os.path.splitext(os.path.basename(bad))[0]): 47}
        results = [run_one(good, mapping, False, None), run_one(bad, mapping, False, None)]
        ready = sum(1 for r in results if r['ok'])
        failed = sum(1 for r in results if not r['ok'])
        self.assertEqual((ready, failed), (1, 1))

    def test_unmapped_driver_is_a_clean_per_file_failure(self):
        good = self._good_file()
        r = run_one(good, {}, commit=False, token=None)
        self.assertFalse(r['ok'])
        self.assertIn('driver-ledger-map.json', r['message'])


class ResolvePathsTests(unittest.TestCase):
    def test_expands_directory_to_sorted_xlsx_files(self):
        d = tempfile.mkdtemp()
        for name in ('b.xlsx', 'a.xlsx', 'notes.txt'):
            open(os.path.join(d, name), 'w').close()
        files = resolve_paths([d])
        self.assertEqual([os.path.basename(f) for f in files], ['a.xlsx', 'b.xlsx'])

    def test_keeps_explicit_files_as_given(self):
        d = tempfile.mkdtemp()
        f1 = os.path.join(d, 'z.xlsx'); f2 = os.path.join(d, 'a.xlsx')
        open(f1, 'w').close(); open(f2, 'w').close()
        self.assertEqual(resolve_paths([f1, f2]), [f1, f2])


if __name__ == '__main__':
    unittest.main()
